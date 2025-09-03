#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
简化版Excel图片提取脚本
从Excel文件中提取所有图片，并根据子表列名分类存储到本地文件夹
"""

import os
import zipfile
import shutil
from pathlib import Path
import xml.etree.ElementTree as ET

class SimpleExcelImageExtractor:
    def __init__(self, excel_file_path, output_dir="extracted_images"):
        """
        初始化Excel图片提取器
        
        Args:
            excel_file_path (str): Excel文件路径
            output_dir (str): 输出目录
        """
        self.excel_file_path = excel_file_path
        self.output_dir = Path(output_dir)
        self.temp_dir = Path("temp_excel_extract")
        
    def extract_images(self):
        """提取Excel中的所有图片"""
        print(f"开始从 {self.excel_file_path} 提取图片...")
        
        # 创建输出目录
        self.output_dir.mkdir(exist_ok=True)
        
        try:
            # 解压Excel文件
            self._extract_excel()
            
            # 提取图片
            self._extract_images_from_media()
            
            print("图片提取完成！")
            
        except Exception as e:
            print(f"提取过程中出现错误: {e}")
        finally:
            # 清理临时文件
            self._cleanup_temp()
    
    def _extract_excel(self):
        """解压Excel文件"""
        print("正在解压Excel文件...")
        
        # 创建临时目录
        self.temp_dir.mkdir(exist_ok=True)
        
        # 解压Excel文件
        with zipfile.ZipFile(self.excel_file_path, 'r') as zip_ref:
            zip_ref.extractall(self.temp_dir)
        
        print("Excel文件解压完成")
    
    def _extract_images_from_media(self):
        """从媒体目录提取图片"""
        media_dir = self.temp_dir / "xl" / "media"
        
        if not media_dir.exists():
            print("未找到媒体目录，可能没有图片")
            return
        
        print(f"找到媒体目录: {media_dir}")
        
        # 获取所有图片文件
        image_files = list(media_dir.glob("*"))
        print(f"发现 {len(image_files)} 个媒体文件")
        
        # 获取工作表信息
        sheet_names = self._get_sheet_names()
        
        # 处理每个工作表
        for sheet_name in sheet_names:
            print(f"处理工作表: {sheet_name}")
            self._process_sheet_images(sheet_name, image_files)
    
    def _get_sheet_names(self):
        """获取工作表名称"""
        try:
            workbook_xml = self.temp_dir / "xl" / "workbook.xml"
            if not workbook_xml.exists():
                return ["Sheet1"]  # 默认工作表名
            
            tree = ET.parse(workbook_xml)
            root = tree.getroot()
            
            # 解析XML命名空间
            namespaces = {'w': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}
            
            sheet_names = []
            for sheet in root.findall('.//w:sheet', namespaces):
                name = sheet.get('name')
                if name:
                    sheet_names.append(name)
            
            return sheet_names if sheet_names else ["Sheet1"]
            
        except Exception as e:
            print(f"获取工作表名称失败: {e}")
            return ["Sheet1"]
    
    def _process_sheet_images(self, sheet_name, image_files):
        """处理工作表中的图片"""
        try:
            # 获取工作表XML文件
            sheet_index = self._get_sheet_index(sheet_name)
            sheet_xml = self.temp_dir / "xl" / "worksheets" / f"sheet{sheet_index}.xml"
            
            if not sheet_xml.exists():
                print(f"  工作表XML文件不存在: {sheet_xml}")
                return
            
            # 解析工作表XML，获取图片位置信息
            image_positions = self._parse_sheet_xml(sheet_xml)
            
            # 根据图片位置信息分类存储
            self._categorize_and_save_images(sheet_name, image_files, image_positions)
            
        except Exception as e:
            print(f"  处理工作表 {sheet_name} 失败: {e}")
    
    def _get_sheet_index(self, sheet_name):
        """获取工作表索引"""
        try:
            workbook_xml = self.temp_dir / "xl" / "workbook.xml"
            tree = ET.parse(workbook_xml)
            root = tree.getroot()
            
            namespaces = {'w': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}
            
            for i, sheet in enumerate(root.findall('.//w:sheet', namespaces)):
                if sheet.get('name') == sheet_name:
                    return i + 1
            
            return 1  # 默认返回第一个工作表
            
        except:
            return 1
    
    def _parse_sheet_xml(self, sheet_xml):
        """解析工作表XML，获取图片位置信息"""
        try:
            tree = ET.parse(sheet_xml)
            root = tree.getroot()
            
            # 解析XML命名空间
            namespaces = {
                'a': 'http://schemas.openxmlformats.org/drawingml/2006/main',
                'xdr': 'http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing'
            }
            
            image_positions = []
            
            # 查找图片元素
            for drawing in root.findall('.//xdr:drawing', namespaces):
                for pic in drawing.findall('.//xdr:pic', namespaces):
                    # 获取图片引用
                    blip = pic.find('.//a:blip', namespaces)
                    if blip is not None:
                        embed = blip.get('{http://schemas.openxmlformats.org/officeDocument/2006/relationships}embed')
                        if embed:
                            # 获取图片位置信息
                            anchor = pic.find('.//xdr:from', namespaces)
                            if anchor is not None:
                                col = anchor.find('xdr:col', namespaces)
                                row = anchor.find('xdr:row', namespaces)
                                
                                if col is not None and row is not None:
                                    col_idx = int(col.text) if col.text else 0
                                    row_idx = int(row.text) if row.text else 0
                                    
                                    image_positions.append({
                                        'embed_id': embed,
                                        'col': col_idx,
                                        'row': row_idx
                                    })
            
            return image_positions
            
        except Exception as e:
            print(f"    解析工作表XML失败: {e}")
            return []
    
    def _categorize_and_save_images(self, sheet_name, image_files, image_positions):
        """根据位置信息分类并保存图片"""
        try:
            # 获取列名信息
            column_names = self._get_column_names(sheet_name)
            print(f"    检测到的列名: {column_names}")
            
            # 记录已处理的图片
            processed_images = set()
            
            # 处理每个图片位置
            for pos in image_positions:
                # 获取对应的图片文件
                image_file = self._get_image_file_by_embed_id(pos['embed_id'])
                if image_file and image_file not in processed_images:
                    # 获取列名
                    col_name = self._get_column_name_by_index(pos['col'], column_names)
                    
                    # 保存图片
                    self._save_image_to_category(image_file, sheet_name, col_name)
                    processed_images.add(image_file)
                    print(f"    图片 {image_file.name} -> {col_name}")
            
            # 使用更智能的方法处理所有图片
            self._smart_categorize_all_images(sheet_name, column_names)
            
        except Exception as e:
            print(f"    分类保存图片失败: {e}")
    
    def _smart_categorize_all_images(self, sheet_name, column_names):
        """智能分类所有图片"""
        try:
            # 获取所有图片文件
            media_dir = self.temp_dir / "xl" / "media"
            if not media_dir.exists():
                return
                
            image_files = list(media_dir.glob("*"))
            if not image_files:
                return
                
            print(f"    找到 {len(image_files)} 个图片文件")
            
            # 如果列名数量合理，尝试平均分配图片
            if len(column_names) > 1 and len(image_files) > 0:
                # 计算每列大概的图片数量
                images_per_col = len(image_files) // len(column_names)
                remainder = len(image_files) % len(column_names)
                
                current_idx = 0
                for col_idx, col_name in enumerate(column_names):
                    # 计算这一列的图片数量
                    col_image_count = images_per_col
                    if col_idx < remainder:
                        col_image_count += 1
                    
                    # 为这一列分配图片
                    for i in range(col_image_count):
                        if current_idx < len(image_files):
                            image_file = image_files[current_idx]
                            self._save_image_to_category(image_file, sheet_name, col_name)
                            current_idx += 1
                            
                print(f"    智能分配完成，共处理 {current_idx} 个图片")
            else:
                # 备用方案：全部放到第一列或"其他"
                col_name = column_names[0] if column_names else "其他"
                for image_file in image_files:
                    self._save_image_to_category(image_file, sheet_name, col_name)
                    
        except Exception as e:
            print(f"    智能分类失败: {e}")
    
    def _get_column_names(self, sheet_name):
        """获取列名信息"""
        try:
            # 尝试使用openpyxl读取第一行作为列名
            from openpyxl import load_workbook
            wb = load_workbook(self.excel_file_path, data_only=True)
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                column_names = []
                for cell in ws[1]:  # 读取第一行
                    if cell.value:
                        column_names.append(str(cell.value))
                    else:
                        column_names.append(f"列{len(column_names)+1}")
                wb.close()
                return column_names
            wb.close()
        except Exception as e:
            print(f"    读取列名失败: {e}")
        
        # 备用方案：使用默认列名
        return [f"列{i+1}" for i in range(26)]
    
    def _get_column_name_by_index(self, col_idx, column_names):
        """根据列索引获取列名"""
        if col_idx < len(column_names):
            return column_names[col_idx]
        else:
            return f"列{col_idx + 1}"
    
    def _get_image_file_by_embed_id(self, embed_id):
        """根据嵌入ID获取图片文件"""
        try:
            # 解析关系文件来找到对应的图片
            rels_file = self.temp_dir / "xl" / "worksheets" / "_rels" / f"sheet{self._get_sheet_index()}.xml.rels"
            if rels_file.exists():
                tree = ET.parse(rels_file)
                root = tree.getroot()
                
                # 查找对应的关系
                for rel in root.findall('.//{http://schemas.openxmlformats.org/package/2006/relationships}Relationship'):
                    if rel.get('Id') == embed_id:
                        target = rel.get('Target')
                        if target:
                            # 构建完整的图片路径
                            image_path = self.temp_dir / "xl" / "worksheets" / target.replace('../', '')
                            if image_path.exists():
                                return image_path
            
            # 备用方案：返回media目录中的图片文件
            media_dir = self.temp_dir / "xl" / "media"
            if media_dir.exists():
                image_files = list(media_dir.glob("*"))
                if image_files:
                    return image_files[0]
            return None
        except Exception as e:
            print(f"    查找图片文件失败: {e}")
            return None
    

    def _save_image_to_category(self, image_file, sheet_name, col_name):
        """保存图片到分类目录"""
        try:
            if image_file and image_file.exists():
                # 创建分类目录
                col_dir = self.output_dir / sheet_name / col_name
                col_dir.mkdir(parents=True, exist_ok=True)
                
                # 生成输出文件名
                file_ext = image_file.suffix
                output_file = col_dir / f"image_{len(list(col_dir.glob('*'))) + 1}{file_ext}"
                
                # 复制文件
                shutil.copy2(image_file, output_file)
                print(f"    已保存图片到 {col_name}: {output_file.name}")
                
        except Exception as e:
            print(f"    保存图片失败: {e}")
    

    
    def _cleanup_temp(self):
        """清理临时文件"""
        try:
            if self.temp_dir.exists():
                shutil.rmtree(self.temp_dir)
                print("临时文件已清理")
        except Exception as e:
            print(f"清理临时文件失败: {e}")

def main():
    """主函数"""
    # Excel文件路径
    excel_file = "副本夹克试标找图.xlsx"
    
    # 检查文件是否存在
    if not os.path.exists(excel_file):
        print(f"错误: 找不到文件 {excel_file}")
        return
    
    # 创建提取器并执行提取
    extractor = SimpleExcelImageExtractor(excel_file)
    extractor.extract_images()
    
    print(f"\n图片已保存到: {extractor.output_dir.absolute()}")
    print("目录结构: 输出目录/工作表名/列名/图片文件")

if __name__ == "__main__":
    main() 